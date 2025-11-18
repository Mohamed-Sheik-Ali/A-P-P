from django.core.management.base import BaseCommand
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import os


class Command(BaseCommand):
    help = 'Create sample Excel template for payroll upload'

    def handle(self, *args, **kwargs):
        # Create workbook
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Payroll Data"
        
        # Headers
        headers = [
            'employee_id', 'name', 'email', 'department', 'designation',
            'basic_pay', 'hra', 'variable_pay', 'special_allowance', 'other_allowances'
        ]
        
        # Style for headers
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        
        # Write headers
        for col_idx, header in enumerate(headers, start=1):
            cell = sheet.cell(row=1, column=col_idx)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Sample data
        sample_data = [
            ['EMP001', 'John Doe', 'john.doe@example.com', 'IT', 'Senior Developer', 50000, 10000, 5000, 2000, 1000],
            ['EMP002', 'Jane Smith', 'jane.smith@example.com', 'HR', 'HR Manager', 60000, 12000, 8000, 3000, 1500],
            ['EMP003', 'Mike Johnson', 'mike.j@example.com', 'Finance', 'Accountant', 45000, 9000, 4000, 1500, 800],
            ['EMP004', 'Sarah Williams', 'sarah.w@example.com', 'Marketing', 'Marketing Lead', 55000, 11000, 6000, 2500, 1200],
            ['EMP005', 'David Brown', 'david.b@example.com', 'IT', 'Junior Developer', 35000, 7000, 3000, 1000, 500],
        ]
        
        # Write sample data
        for row_idx, row_data in enumerate(sample_data, start=2):
            for col_idx, value in enumerate(row_data, start=1):
                sheet.cell(row=row_idx, column=col_idx).value = value
        
        # Auto-adjust column widths
        for column in sheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            sheet.column_dimensions[column_letter].width = adjusted_width
        
        # Save file
        filename = 'sample_payroll_template.xlsx'
        workbook.save(filename)
        
        self.stdout.write(self.style.SUCCESS(f'Successfully created {filename}'))