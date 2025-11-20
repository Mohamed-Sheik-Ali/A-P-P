import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from decimal import Decimal
from django.core.files.base import ContentFile
from django.utils import timezone
from io import BytesIO
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
import os

from .models import PayrollUpload, Employee, SalaryComponent, PayrollReport


class ExcelProcessor:
    """Class to handle Excel file processing"""
    
    # Expected column names (case-insensitive)
    EXPECTED_COLUMNS = [
        'employee_id', 'name', 'email', 'department', 'designation',
        'basic_pay', 'hra', 'variable_pay', 'special_allowance', 'other_allowances'
    ]
    
    def __init__(self, file_path, upload_instance):
        self.file_path = file_path
        self.upload_instance = upload_instance
        self.errors = []
        self.warnings = []
    
    def validate_file(self):
        """Validate Excel file structure"""
        try:
            workbook = openpyxl.load_workbook(self.file_path)
            sheet = workbook.active
            
            # Check if file is empty
            if sheet.max_row < 2:
                self.errors.append("Excel file is empty or has no data rows")
                return False
            
            # Get header row
            headers = [cell.value for cell in sheet[1]]
            headers = [str(h).strip().lower().replace(' ', '_') if h else '' for h in headers]
            
            # Check for required columns
            missing_columns = []
            for col in self.EXPECTED_COLUMNS:
                if col not in headers:
                    missing_columns.append(col)
            
            if missing_columns:
                self.errors.append(f"Missing required columns: {', '.join(missing_columns)}")
                return False
            
            # Check row limit
            if sheet.max_row > 101:  # 1 header + 100 data rows
                self.errors.append("Excel file contains more than 100 employees. Maximum limit is 100.")
                return False
            
            return True
            
        except Exception as e:
            self.errors.append(f"Error reading Excel file: {str(e)}")
            return False
    
    def parse_and_save(self):
        """Parse Excel file and save to database"""
        try:
            workbook = openpyxl.load_workbook(self.file_path)
            sheet = workbook.active
            
            # Get headers
            headers = [cell.value for cell in sheet[1]]
            headers = [str(h).strip().lower().replace(' ', '_') if h else '' for h in headers]
            
            # Create column mapping
            col_map = {header: idx for idx, header in enumerate(headers)}
            
            employees_created = 0
            
            # Process each row
            for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                try:
                    # Extract employee data
                    employee_id = str(row[col_map['employee_id']]).strip() if row[col_map['employee_id']] else None
                    name = str(row[col_map['name']]).strip() if row[col_map['name']] else None
                    
                    # Skip empty rows
                    if not employee_id or not name:
                        self.warnings.append(f"Row {row_idx}: Skipped empty row")
                        continue
                    
                    email = str(row[col_map['email']]).strip() if row[col_map.get('email')] and row[col_map.get('email')] else None
                    department = str(row[col_map['department']]).strip() if row[col_map.get('department')] and row[col_map.get('department')] else None
                    designation = str(row[col_map['designation']]).strip() if row[col_map.get('designation')] and row[col_map.get('designation')] else None
                    
                    # Create employee
                    employee = Employee.objects.create(
                        upload=self.upload_instance,
                        employee_id=employee_id,
                        name=name,
                        email=email,
                        department=department,
                        designation=designation
                    )
                    
                    # Extract salary components
                    def get_decimal_value(value):
                        if value is None or value == '':
                            return Decimal('0.00')
                        try:
                            return Decimal(str(value))
                        except:
                            return Decimal('0.00')
                    
                    basic_pay = get_decimal_value(row[col_map['basic_pay']])
                    hra = get_decimal_value(row[col_map['hra']])
                    variable_pay = get_decimal_value(row[col_map['variable_pay']])
                    special_allowance = get_decimal_value(row[col_map.get('special_allowance', 0)])
                    other_allowances = get_decimal_value(row[col_map.get('other_allowances', 0)])
                    
                    # Create salary component
                    salary = SalaryComponent.objects.create(
                        employee=employee,
                        basic_pay=basic_pay,
                        hra=hra,
                        variable_pay=variable_pay,
                        special_allowance=special_allowance,
                        other_allowances=other_allowances
                    )
                    
                    # Calculate salary
                    salary.calculate_salary()
                    
                    employees_created += 1
                    
                except Exception as e:
                    self.errors.append(f"Row {row_idx}: Error processing employee - {str(e)}")
                    continue
            
            # Update upload instance
            self.upload_instance.total_employees = employees_created
            self.upload_instance.status = 'completed'
            self.upload_instance.processed_date = timezone.now()
            self.upload_instance.save()
            
            return True, employees_created
            
        except Exception as e:
            self.errors.append(f"Error parsing Excel file: {str(e)}")
            self.upload_instance.status = 'failed'
            self.upload_instance.error_message = str(e)
            self.upload_instance.save()
            return False, 0


class ReportGenerator:
    """Class to generate payroll reports"""
    
    def __init__(self, upload_instance):
        self.upload_instance = upload_instance
        # Get employees who have salary components in this upload
        employee_ids = upload_instance.salary_components.values_list('employee_id', flat=True).distinct()
        self.employees = Employee.objects.filter(id__in=employee_ids)
        self.upload_id = upload_instance.id
    
    def generate_excel_report(self):
        """Generate Excel report"""
        try:
            # Create workbook
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = "Payroll Report"
            
            # Define styles
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=11)
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Headers
            headers = [
                'Employee ID', 'Name', 'Email', 'Department', 'Designation',
                'Basic Pay', 'HRA', 'Variable Pay', 'Special Allowance', 'Other Allowances',
                'Gross Salary', 'PF', 'Professional Tax', 'Income Tax', 'Other Deductions',
                'Total Deductions', 'Net Salary', 'Take Home Pay'
            ]
            
            # Write headers
            for col_idx, header in enumerate(headers, start=1):
                cell = sheet.cell(row=1, column=col_idx)
                cell.value = header
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = border
            
            # Write data
            for row_idx, employee in enumerate(self.employees, start=2):
                # Get salary component for this employee and upload
                try:
                    salary = SalaryComponent.objects.get(employee=employee, upload_id=self.upload_id)
                except SalaryComponent.DoesNotExist:
                    continue  # Skip employees without salary data
                
                data = [
                    employee.employee_id,
                    employee.name,
                    employee.email or '',
                    employee.department or '',
                    employee.designation or '',
                    float(salary.basic_pay),
                    float(salary.hra),
                    float(salary.variable_pay),
                    float(salary.special_allowance),
                    float(salary.other_allowances),
                    float(salary.gross_salary),
                    float(salary.provident_fund),
                    float(salary.professional_tax),
                    float(salary.income_tax),
                    float(salary.other_deductions),
                    float(salary.total_deductions),
                    float(salary.net_salary),
                    float(salary.take_home_pay)
                ]
                
                for col_idx, value in enumerate(data, start=1):
                    cell = sheet.cell(row=row_idx, column=col_idx)
                    cell.value = value
                    cell.border = border
                    
                    # Format currency columns
                    if col_idx >= 6:
                        cell.number_format = '₹#,##0.00'
                        cell.alignment = Alignment(horizontal='right')
                    else:
                        cell.alignment = Alignment(horizontal='left')
            
            # Auto-adjust column widths
            for column in sheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                sheet.column_dimensions[column_letter].width = adjusted_width
            
            # Add summary section
            summary_row = len(self.employees) + 3
            sheet.cell(row=summary_row, column=1).value = "SUMMARY"
            sheet.cell(row=summary_row, column=1).font = Font(bold=True, size=12)
            
            # Calculate totals
            total_gross = 0
            total_deductions = 0
            total_net = 0
            
            for employee in self.employees:
                try:
                    salary = SalaryComponent.objects.get(employee=employee, upload_id=self.upload_id)
                    total_gross += salary.gross_salary
                    total_deductions += salary.total_deductions
                    total_net += salary.net_salary
                except SalaryComponent.DoesNotExist:
                    continue
            
            summary_data = [
                ("Total Employees", len(self.employees)),
                ("Total Gross Salary", float(total_gross)),
                ("Total Deductions", float(total_deductions)),
                ("Total Net Salary", float(total_net))
            ]
            
            for idx, (label, value) in enumerate(summary_data, start=1):
                sheet.cell(row=summary_row + idx, column=1).value = label
                sheet.cell(row=summary_row + idx, column=1).font = Font(bold=True)
                cell = sheet.cell(row=summary_row + idx, column=2)
                cell.value = value
                if isinstance(value, (int, float)) and label != "Total Employees":
                    cell.number_format = '₹#,##0.00'
            
            # Save to BytesIO
            excel_file = BytesIO()
            workbook.save(excel_file)
            excel_file.seek(0)
            
            # Create PayrollReport record
            filename = f"payroll_report_{self.upload_instance.id}_{timezone.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            
            report = PayrollReport.objects.create(
                upload=self.upload_instance,
                report_type='excel',
                file_size=excel_file.getbuffer().nbytes
            )
            
            report.file.save(filename, ContentFile(excel_file.read()), save=True)
            
            return report
            
        except Exception as e:
            raise Exception(f"Error generating Excel report: {str(e)}")
    
    def generate_pdf_report(self):
        """Generate PDF report"""
        try:
            # Create PDF buffer
            pdf_buffer = BytesIO()
            
            # Create document
            doc = SimpleDocTemplate(
                pdf_buffer,
                pagesize=A4,
                rightMargin=30,
                leftMargin=30,
                topMargin=30,
                bottomMargin=30
            )
            
            # Container for elements
            elements = []
            
            # Styles
            styles = getSampleStyleSheet()
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=24,
                textColor=colors.HexColor('#4472C4'),
                spaceAfter=30,
                alignment=TA_CENTER
            )
            
            heading_style = ParagraphStyle(
                'CustomHeading',
                parent=styles['Heading2'],
                fontSize=14,
                textColor=colors.HexColor('#4472C4'),
                spaceAfter=12,
                spaceBefore=12
            )
            
            # Title
            title = Paragraph("Payroll Report", title_style)
            elements.append(title)
            
            # Report info
            info_data = [
                ['Report Date:', timezone.now().strftime('%B %d, %Y')],
                ['Upload ID:', str(self.upload_instance.id)],
                ['Total Employees:', str(len(self.employees))],
                ['Processed By:', self.upload_instance.user.get_full_name() or self.upload_instance.user.username]
            ]
            
            info_table = Table(info_data, colWidths=[2*inch, 4*inch])
            info_table.setStyle(TableStyle([
                ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 10),
                ('TEXTCOLOR', (0, 0), (0, -1), colors.HexColor('#4472C4')),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
            ]))
            
            elements.append(info_table)
            elements.append(Spacer(1, 20))
            
            # Employee details
            for employee in self.employees:
                # Get salary component for this employee and upload
                try:
                    salary = SalaryComponent.objects.get(employee=employee, upload_id=self.upload_id)
                except SalaryComponent.DoesNotExist:
                    continue  # Skip employees without salary data
                
                # Employee header
                emp_heading = Paragraph(f"{employee.name} ({employee.employee_id})", heading_style)
                elements.append(emp_heading)
                
                # Employee info
                emp_info = []
                if employee.email:
                    emp_info.append(['Email:', employee.email])
                if employee.department:
                    emp_info.append(['Department:', employee.department])
                if employee.designation:
                    emp_info.append(['Designation:', employee.designation])
                
                if emp_info:
                    emp_info_table = Table(emp_info, colWidths=[1.5*inch, 4*inch])
                    emp_info_table.setStyle(TableStyle([
                        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
                        ('FONTSIZE', (0, 0), (-1, -1), 9),
                        ('TEXTCOLOR', (0, 0), (0, -1), colors.grey),
                        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
                    ]))
                    elements.append(emp_info_table)
                    elements.append(Spacer(1, 10))
                
                # Salary breakdown
                salary_data = [
                    ['Component', 'Amount (₹)'],
                    ['Basic Pay', f'{salary.basic_pay:,.2f}'],
                    ['HRA', f'{salary.hra:,.2f}'],
                    ['Variable Pay', f'{salary.variable_pay:,.2f}'],
                    ['Special Allowance', f'{salary.special_allowance:,.2f}'],
                    ['Other Allowances', f'{salary.other_allowances:,.2f}'],
                    ['Gross Salary', f'{salary.gross_salary:,.2f}'],
                    ['', ''],
                    ['Provident Fund', f'{salary.provident_fund:,.2f}'],
                    ['Professional Tax', f'{salary.professional_tax:,.2f}'],
                    ['Income Tax', f'{salary.income_tax:,.2f}'],
                    ['Other Deductions', f'{salary.other_deductions:,.2f}'],
                    ['Total Deductions', f'{salary.total_deductions:,.2f}'],
                    ['', ''],
                    ['Net Salary', f'{salary.net_salary:,.2f}'],
                    ['Take Home Pay', f'{salary.take_home_pay:,.2f}'],
                ]
                
                salary_table = Table(salary_data, colWidths=[3*inch, 2*inch])
                salary_table.setStyle(TableStyle([
                    # Header
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 10),
                    ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
                    
                    # Body
                    ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                    ('FONTSIZE', (0, 1), (-1, -1), 9),
                    ('ALIGN', (1, 1), (1, -1), 'RIGHT'),
                    
                    # Gross salary row
                    ('BACKGROUND', (0, 6), (-1, 6), colors.HexColor('#E7E6E6')),
                    ('FONTNAME', (0, 6), (-1, 6), 'Helvetica-Bold'),
                    
                    # Total deductions row
                    ('BACKGROUND', (0, 12), (-1, 12), colors.HexColor('#E7E6E6')),
                    ('FONTNAME', (0, 12), (-1, 12), 'Helvetica-Bold'),
                    
                    # Final rows
                    ('BACKGROUND', (0, 14), (-1, -1), colors.HexColor('#4472C4')),
                    ('TEXTCOLOR', (0, 14), (-1, -1), colors.whitesmoke),
                    ('FONTNAME', (0, 14), (-1, -1), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 14), (-1, -1), 10),
                    
                    # Grid
                    ('GRID', (0, 0), (-1, -1), 1, colors.grey),
                    ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                    ('LEFTPADDING', (0, 0), (-1, -1), 8),
                    ('RIGHTPADDING', (0, 0), (-1, -1), 8),
                    ('TOPPADDING', (0, 0), (-1, -1), 6),
                    ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
                ]))
                
                elements.append(salary_table)
                elements.append(PageBreak())
            
            # Summary page
            summary_heading = Paragraph("Summary", title_style)
            elements.append(summary_heading)
            
            total_gross = 0
            total_deductions = 0
            total_net = 0
            
            for employee in self.employees:
                try:
                    salary = SalaryComponent.objects.get(employee=employee, upload_id=self.upload_id)
                    total_gross += salary.gross_salary
                    total_deductions += salary.total_deductions
                    total_net += salary.net_salary
                except SalaryComponent.DoesNotExist:
                    continue
            
            summary_data = [
                ['Metric', 'Value'],
                ['Total Employees', str(len(self.employees))],
                ['Total Gross Salary', f'₹{total_gross:,.2f}'],
                ['Total Deductions', f'₹{total_deductions:,.2f}'],
                ['Total Net Salary', f'₹{total_net:,.2f}'],
                ['Average Gross Salary', f'₹{total_gross/len(self.employees):,.2f}'],
                ['Average Net Salary', f'₹{total_net/len(self.employees):,.2f}'],
            ]
            
            summary_table = Table(summary_data, colWidths=[3*inch, 3*inch])
            summary_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 1), (-1, -1), 11),
                ('ALIGN', (1, 1), (1, -1), 'RIGHT'),
                ('GRID', (0, 0), (-1, -1), 1, colors.grey),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('LEFTPADDING', (0, 0), (-1, -1), 12),
                ('RIGHTPADDING', (0, 0), (-1, -1), 12),
                ('TOPPADDING', (0, 0), (-1, -1), 10),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 10),
            ]))
            
            elements.append(summary_table)
            
            # Build PDF
            doc.build(elements)
            
            # Get PDF content
            pdf_buffer.seek(0)
            pdf_content = pdf_buffer.read()
            
            # Create PayrollReport record
            filename = f"payroll_report_{self.upload_instance.id}_{timezone.now().strftime('%Y%m%d_%H%M%S')}.pdf"
            
            report = PayrollReport.objects.create(
                upload=self.upload_instance,
                report_type='pdf',
                file_size=len(pdf_content)
            )
            
            report.file.save(filename, ContentFile(pdf_content), save=True)
            
            return report
            
        except Exception as e:
            raise Exception(f"Error generating PDF report: {str(e)}")


class IndividualEmployeeReportGenerator:
    """Class to generate individual employee reports"""
    
    def __init__(self, employee, upload_id=None):
        self.employee = employee
        self.upload_id = upload_id
        
        # Get upload instance if upload_id is provided
        if upload_id:
            self.upload_instance = PayrollUpload.objects.get(id=upload_id)
        else:
            self.upload_instance = None
            
        # Get the latest salary component for this employee
        if upload_id:
            try:
                self.salary = SalaryComponent.objects.get(employee=employee, upload_id=upload_id)
            except SalaryComponent.DoesNotExist:
                # Fall back to latest salary record
                self.salary = employee.salary_records.first()
        else:
            # Get latest salary record
            self.salary = employee.salary_records.first()
        
        if not self.salary:
            raise ValueError(f"No salary data found for employee {employee.name}")
    
    def generate_excel_report(self):
        """Generate Excel report for individual employee"""
        try:
            # Create workbook
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = f"{self.employee.name} - Payroll"
            
            # Define styles
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=12)
            section_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
            section_font = Font(bold=True, size=11)
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Title
            sheet.merge_cells('A1:D1')
            title_cell = sheet.cell(row=1, column=1)
            title_cell.value = f"PAYROLL SLIP - {self.employee.name.upper()}"
            title_cell.fill = header_fill
            title_cell.font = header_font
            title_cell.alignment = Alignment(horizontal='center', vertical='center')
            title_cell.border = border
            
            # Employee Information Section
            row = 3
            sheet.merge_cells(f'A{row}:D{row}')
            info_header = sheet.cell(row=row, column=1)
            info_header.value = "EMPLOYEE INFORMATION"
            info_header.fill = section_fill
            info_header.font = section_font
            info_header.border = border
            
            employee_info = [
                ['Employee ID', self.employee.employee_id, 'Name', self.employee.name],
                ['Department', self.employee.department or 'N/A', 'Designation', self.employee.designation or 'N/A'],
                ['Email', self.employee.email or 'N/A', 'Upload Date', self.upload_instance.upload_date.strftime('%Y-%m-%d') if self.upload_instance else 'N/A']
            ]
            
            row += 1
            for info_row in employee_info:
                for col_idx, value in enumerate(info_row, start=1):
                    cell = sheet.cell(row=row, column=col_idx)
                    cell.value = value
                    if col_idx in [1, 3]:  # Labels
                        cell.font = Font(bold=True)
                    cell.border = border
                    cell.alignment = Alignment(horizontal='center' if col_idx in [1, 3] else 'left')
                row += 1
            
            # Earnings Section
            row += 1
            sheet.merge_cells(f'A{row}:D{row}')
            earnings_header = sheet.cell(row=row, column=1)
            earnings_header.value = "EARNINGS"
            earnings_header.fill = section_fill
            earnings_header.font = section_font
            earnings_header.border = border
            
            earnings_data = [
                ['Component', 'Amount', '', ''],
                ['Basic Pay', f'₹{self.salary.basic_pay:,.2f}', '', ''],
                ['HRA', f'₹{self.salary.hra:,.2f}', '', ''],
                ['Variable Pay', f'₹{self.salary.variable_pay:,.2f}', '', ''],
                ['Special Allowance', f'₹{self.salary.special_allowance:,.2f}', '', ''],
                ['Other Allowances', f'₹{self.salary.other_allowances:,.2f}', '', ''],
                ['GROSS SALARY', f'₹{self.salary.gross_salary:,.2f}', '', ''],
            ]
            
            row += 1
            for earn_row in earnings_data:
                for col_idx, value in enumerate(earn_row, start=1):
                    if col_idx <= 2:  # Only use first 2 columns
                        cell = sheet.cell(row=row, column=col_idx)
                        cell.value = value
                        if 'GROSS SALARY' in value or 'Component' in value:
                            cell.font = Font(bold=True)
                            if 'GROSS SALARY' in value:
                                cell.fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
                        cell.border = border
                        cell.alignment = Alignment(horizontal='left' if col_idx == 1 else 'right')
                row += 1
            
            # Deductions Section
            row += 1
            sheet.merge_cells(f'A{row}:D{row}')
            deductions_header = sheet.cell(row=row, column=1)
            deductions_header.value = "DEDUCTIONS"
            deductions_header.fill = section_fill
            deductions_header.font = section_font
            deductions_header.border = border
            
            deductions_data = [
                ['Component', 'Amount', '', ''],
                ['Provident Fund (12%)', f'₹{self.salary.provident_fund:,.2f}', '', ''],
                ['Professional Tax', f'₹{self.salary.professional_tax:,.2f}', '', ''],
                ['Income Tax', f'₹{self.salary.income_tax:,.2f}', '', ''],
                ['Other Deductions', f'₹{self.salary.other_deductions:,.2f}', '', ''],
                ['TOTAL DEDUCTIONS', f'₹{self.salary.total_deductions:,.2f}', '', ''],
            ]
            
            row += 1
            for deduct_row in deductions_data:
                for col_idx, value in enumerate(deduct_row, start=1):
                    if col_idx <= 2:  # Only use first 2 columns
                        cell = sheet.cell(row=row, column=col_idx)
                        cell.value = value
                        if 'TOTAL DEDUCTIONS' in value or 'Component' in value:
                            cell.font = Font(bold=True)
                            if 'TOTAL DEDUCTIONS' in value:
                                cell.fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
                        cell.border = border
                        cell.alignment = Alignment(horizontal='left' if col_idx == 1 else 'right')
                row += 1
            
            # Net Salary Section
            row += 1
            sheet.merge_cells(f'A{row}:B{row}')
            net_cell = sheet.cell(row=row, column=1)
            net_cell.value = f"NET SALARY: ₹{self.salary.net_salary:,.2f}"
            net_cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
            net_cell.font = Font(bold=True, color="FFFFFF", size=14)
            net_cell.alignment = Alignment(horizontal='center', vertical='center')
            net_cell.border = border
            
            # Adjust column widths
            sheet.column_dimensions['A'].width = 20
            sheet.column_dimensions['B'].width = 15
            sheet.column_dimensions['C'].width = 15
            sheet.column_dimensions['D'].width = 20
            
            # Save to BytesIO
            excel_buffer = BytesIO()
            workbook.save(excel_buffer)
            excel_buffer.seek(0)
            excel_content = excel_buffer.read()
            
            filename = f"{self.employee.name.replace(' ', '_')}_payroll_{timezone.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            content_type = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            
            return excel_content, filename, content_type
            
        except Exception as e:
            raise Exception(f"Error generating Excel report: {str(e)}")
    
    def generate_pdf_report(self):
        """Generate PDF report for individual employee"""
        try:
            # Create PDF buffer
            pdf_buffer = BytesIO()
            doc = SimpleDocTemplate(pdf_buffer, pagesize=letter, 
                                  rightMargin=50, leftMargin=50, 
                                  topMargin=50, bottomMargin=50)
            
            # Get styles
            styles = getSampleStyleSheet()
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=16,
                spaceAfter=30,
                alignment=TA_CENTER,
                textColor=colors.HexColor('#4472C4')
            )
            
            heading_style = ParagraphStyle(
                'CustomHeading',
                parent=styles['Heading2'],
                fontSize=12,
                spaceAfter=12,
                textColor=colors.HexColor('#2F5597')
            )
            
            # Build PDF elements
            elements = []
            
            # Title
            title = Paragraph(f"PAYROLL SLIP - {self.employee.name.upper()}", title_style)
            elements.append(title)
            
            # Employee Information
            emp_heading = Paragraph("Employee Information", heading_style)
            elements.append(emp_heading)
            
            emp_data = [
                ['Employee ID', self.employee.employee_id],
                ['Name', self.employee.name],
                ['Department', self.employee.department or 'N/A'],
                ['Designation', self.employee.designation or 'N/A'],
                ['Email', self.employee.email or 'N/A'],
                ['Pay Period', self.upload_instance.upload_date.strftime('%B %Y') if self.upload_instance else 'N/A']
            ]
            
            emp_table = Table(emp_data, colWidths=[2*inch, 3*inch])
            emp_table.setStyle(TableStyle([
                ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
                ('FONTNAME', (1, 0), (1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 0), (-1, -1), 11),
                ('GRID', (0, 0), (-1, -1), 1, colors.grey),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('LEFTPADDING', (0, 0), (-1, -1), 10),
                ('RIGHTPADDING', (0, 0), (-1, -1), 10),
                ('TOPPADDING', (0, 0), (-1, -1), 8),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
            ]))
            
            elements.append(emp_table)
            elements.append(Spacer(1, 20))
            
            # Earnings
            earnings_heading = Paragraph("Earnings", heading_style)
            elements.append(earnings_heading)
            
            earnings_data = [
                ['Component', 'Amount'],
                ['Basic Pay', f'₹{self.salary.basic_pay:,.2f}'],
                ['HRA', f'₹{self.salary.hra:,.2f}'],
                ['Variable Pay', f'₹{self.salary.variable_pay:,.2f}'],
                ['Special Allowance', f'₹{self.salary.special_allowance:,.2f}'],
                ['Other Allowances', f'₹{self.salary.other_allowances:,.2f}'],
                ['GROSS SALARY', f'₹{self.salary.gross_salary:,.2f}']
            ]
            
            earnings_table = Table(earnings_data, colWidths=[2.5*inch, 2.5*inch])
            earnings_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 11),
                ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
                ('FONTNAME', (0, 1), (-1, -2), 'Helvetica'),
                ('FONTSIZE', (0, 1), (-1, -2), 10),
                ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
                ('FONTSIZE', (0, -1), (-1, -1), 11),
                ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#E2EFDA')),
                ('ALIGN', (1, 1), (1, -1), 'RIGHT'),
                ('GRID', (0, 0), (-1, -1), 1, colors.grey),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('LEFTPADDING', (0, 0), (-1, -1), 10),
                ('RIGHTPADDING', (0, 0), (-1, -1), 10),
                ('TOPPADDING', (0, 0), (-1, -1), 8),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
            ]))
            
            elements.append(earnings_table)
            elements.append(Spacer(1, 20))
            
            # Deductions
            deductions_heading = Paragraph("Deductions", heading_style)
            elements.append(deductions_heading)
            
            deductions_data = [
                ['Component', 'Amount'],
                ['Provident Fund (12%)', f'₹{self.salary.provident_fund:,.2f}'],
                ['Professional Tax', f'₹{self.salary.professional_tax:,.2f}'],
                ['Income Tax', f'₹{self.salary.income_tax:,.2f}'],
                ['Other Deductions', f'₹{self.salary.other_deductions:,.2f}'],
                ['TOTAL DEDUCTIONS', f'₹{self.salary.total_deductions:,.2f}']
            ]
            
            deductions_table = Table(deductions_data, colWidths=[2.5*inch, 2.5*inch])
            deductions_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 11),
                ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
                ('FONTNAME', (0, 1), (-1, -2), 'Helvetica'),
                ('FONTSIZE', (0, 1), (-1, -2), 10),
                ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
                ('FONTSIZE', (0, -1), (-1, -1), 11),
                ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#FCE4D6')),
                ('ALIGN', (1, 1), (1, -1), 'RIGHT'),
                ('GRID', (0, 0), (-1, -1), 1, colors.grey),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('LEFTPADDING', (0, 0), (-1, -1), 10),
                ('RIGHTPADDING', (0, 0), (-1, -1), 10),
                ('TOPPADDING', (0, 0), (-1, -1), 8),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
            ]))
            
            elements.append(deductions_table)
            elements.append(Spacer(1, 30))
            
            # Net Salary (Highlighted)
            net_salary_text = f"<b>NET SALARY: ₹{self.salary.net_salary:,.2f}</b>"
            net_style = ParagraphStyle(
                'NetSalary',
                parent=styles['Normal'],
                fontSize=16,
                alignment=TA_CENTER,
                textColor=colors.HexColor('#70AD47'),
                borderColor=colors.HexColor('#70AD47'),
                borderWidth=2,
                borderPadding=15,
                spaceAfter=20
            )
            
            net_paragraph = Paragraph(net_salary_text, net_style)
            elements.append(net_paragraph)
            
            # Footer with generation info
            footer_text = f"Generated on: {timezone.now().strftime('%B %d, %Y at %I:%M %p')}"
            footer_style = ParagraphStyle(
                'Footer',
                parent=styles['Normal'],
                fontSize=8,
                alignment=TA_CENTER,
                textColor=colors.grey
            )
            footer = Paragraph(footer_text, footer_style)
            elements.append(footer)
            
            # Build PDF
            doc.build(elements)
            
            # Get PDF content
            pdf_buffer.seek(0)
            pdf_content = pdf_buffer.read()
            
            filename = f"{self.employee.name.replace(' ', '_')}_payroll_{timezone.now().strftime('%Y%m%d_%H%M%S')}.pdf"
            content_type = 'application/pdf'
            
            return pdf_content, filename, content_type
            
        except Exception as e:
            raise Exception(f"Error generating PDF report: {str(e)}")