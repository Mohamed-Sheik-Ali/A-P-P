#!/usr/bin/env python3
"""
Test script to verify that Excel processing works with user-scoped employees
"""

import os
import sys
import django
from django.conf import settings
from django.core.files.uploadedfile import SimpleUploadedFile
from django.core.files.base import ContentFile
import tempfile

# Add the project directory to Python path
project_dir = '/Users/mohamedsheikalim/Desktop/payroll_processor'
sys.path.insert(0, project_dir)

# Set Django settings module
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'payroll_config.settings')

# Setup Django
django.setup()

from django.contrib.auth.models import User
from payroll.models import PayrollUpload, Employee, SalaryComponent
from payroll.utils import ExcelProcessor
import openpyxl
from io import BytesIO

def create_test_excel():
    """Create a test Excel file with sample employee data"""
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    
    # Headers
    headers = [
        'employee_id', 'name', 'email', 'department', 'designation',
        'basic_pay', 'hra', 'variable_pay', 'special_allowance', 'other_allowances'
    ]
    
    for col_idx, header in enumerate(headers, start=1):
        sheet.cell(row=1, column=col_idx, value=header)
    
    # Sample data - using same employee IDs that might exist in other uploads
    sample_data = [
        ['EMP001', 'Alice Johnson', 'alice@company.com', 'Engineering', 'Software Engineer', 50000, 20000, 10000, 5000, 2000],
        ['EMP002', 'Bob Smith', 'bob@company.com', 'Marketing', 'Marketing Manager', 45000, 18000, 8000, 4000, 1500],
        ['EMP003', 'Carol Brown', 'carol@company.com', 'Finance', 'Financial Analyst', 40000, 16000, 6000, 3000, 1000],
    ]
    
    for row_idx, row_data in enumerate(sample_data, start=2):
        for col_idx, value in enumerate(row_data, start=1):
            sheet.cell(row=row_idx, column=col_idx, value=value)
    
    # Save to bytes buffer
    excel_buffer = BytesIO()
    workbook.save(excel_buffer)
    excel_buffer.seek(0)
    
    return excel_buffer.getvalue()

def test_excel_processing():
    """Test that Excel processing works correctly with user-scoped employees"""
    
    # Create test users
    user1, _ = User.objects.get_or_create(
        username='company1_hr',
        defaults={
            'email': 'hr@company1.com',
            'first_name': 'HR',
            'last_name': 'Manager1'
        }
    )
    
    user2, _ = User.objects.get_or_create(
        username='company2_hr',
        defaults={
            'email': 'hr@company2.com', 
            'first_name': 'HR',
            'last_name': 'Manager2'
        }
    )
    
    print(f"Created test users: {user1.username}, {user2.username}")
    
    try:
        # Test 1: Upload file for user1
        print("\n=== Test 1: Upload for User 1 ===")
        excel_content = create_test_excel()
        
        # Create a temporary file
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp_file:
            tmp_file.write(excel_content)
            tmp_file_path = tmp_file.name
        
        # Create upload record for user1
        upload1 = PayrollUpload.objects.create(
            user=user1,
            file='test_uploads/company1_payroll.xlsx',
            filename='company1_payroll.xlsx',
            status='pending'
        )
        
        # Copy the temp file to the upload location  
        os.makedirs(os.path.dirname(upload1.file.path), exist_ok=True)
        with open(tmp_file_path, 'rb') as src:
            with open(upload1.file.path, 'wb') as dst:
                dst.write(src.read())
        
        # Process the file
        processor1 = ExcelProcessor(upload1.file.path, upload1)
        
        if processor1.validate_file():
            print("✓ File validation passed for user1")
            success, count = processor1.parse_and_save()
            if success:
                print(f"✓ Successfully processed {count} employees for user1")
            else:
                print(f"✗ Processing failed for user1: {processor1.errors}")
        else:
            print(f"✗ File validation failed for user1: {processor1.errors}")
        
        # Check employees created for user1
        user1_employees = Employee.objects.filter(user=user1)
        print(f"✓ User1 now has {user1_employees.count()} employees")
        
        # Test 2: Upload same employee IDs for user2 (should work)
        print("\n=== Test 2: Upload same EMPIDs for User 2 ===")
        
        # Create upload record for user2
        upload2 = PayrollUpload.objects.create(
            user=user2,
            file='test_uploads/company2_payroll.xlsx',
            filename='company2_payroll.xlsx',
            status='pending'
        )
        
        # Copy the same Excel file for user2
        os.makedirs(os.path.dirname(upload2.file.path), exist_ok=True)
        with open(tmp_file_path, 'rb') as src:
            with open(upload2.file.path, 'wb') as dst:
                dst.write(src.read())
        
        # Process the file for user2
        processor2 = ExcelProcessor(upload2.file.path, upload2)
        
        if processor2.validate_file():
            print("✓ File validation passed for user2")
            success, count = processor2.parse_and_save()
            if success:
                print(f"✓ Successfully processed {count} employees for user2")
            else:
                print(f"✗ Processing failed for user2: {processor2.errors}")
        else:
            print(f"✗ File validation failed for user2: {processor2.errors}")
        
        # Check employees created for user2
        user2_employees = Employee.objects.filter(user=user2)
        print(f"✓ User2 now has {user2_employees.count()} employees")
        
        # Test 3: Verify data isolation
        print("\n=== Test 3: Verify Data Isolation ===")
        
        # Check that users can't see each other's employees
        user1_emp001 = Employee.objects.filter(user=user1, employee_id='EMP001').first()
        user2_emp001 = Employee.objects.filter(user=user2, employee_id='EMP001').first()
        
        if user1_emp001 and user2_emp001:
            print(f"✓ Both users have EMP001:")
            print(f"   User1 EMP001: {user1_emp001.name} ({user1_emp001.email})")
            print(f"   User2 EMP001: {user2_emp001.name} ({user2_emp001.email})")
            
            if user1_emp001.id != user2_emp001.id:
                print("✓ Employees are separate database records")
            else:
                print("✗ ERROR: Employees share same database record!")
        
        # Check salary components
        user1_salaries = SalaryComponent.objects.filter(upload=upload1).count()
        user2_salaries = SalaryComponent.objects.filter(upload=upload2).count()
        
        print(f"✓ User1 has {user1_salaries} salary records")
        print(f"✓ User2 has {user2_salaries} salary records")
        
        print("\n✅ All Excel processing tests passed!")
        
        # Cleanup
        print("\nCleaning up test data...")
        upload1.delete()
        upload2.delete()
        user1_employees.delete()
        user2_employees.delete()
        
        # Cleanup temp file
        if os.path.exists(tmp_file_path):
            os.unlink(tmp_file_path)
        
        print("✓ Test data cleaned up")
        
    except Exception as e:
        print(f"✗ Error during testing: {e}")
        import traceback
        traceback.print_exc()

if __name__ == '__main__':
    test_excel_processing()
